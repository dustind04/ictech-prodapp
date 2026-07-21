"""
Weekly Input List importer.

Dave produces "<date> Input List.xlsx" every week (sheets: 'Input List',
'MyMix', 'Mic Assign'). This module turns that workbook into a
slot-assignment plan:

  parse_workbook(bytes)  -> what the spreadsheet says
  build_plan(db, parsed) -> resolved against people/channels/positions,
                            with per-slot warnings for anything unmatched
  apply_plan(db, plan)   -> writes the assignments

The flow is deliberately two-step (preview, then apply) — the admin sees
exactly what will land on the wall before it does.

Matching philosophy: the spreadsheet is Dave's language ("Chris Vox",
"Red HH", "Pack 1", "Pool 2"). We match loosely — first names against
people, color tags against channel labels, pack numbers against IEM
labels — and warn rather than guess when a match isn't clean. A slot
row with no matching person is still applied (empty slot): the sheet is
the weekly truth, including who ISN'T on this week.
"""

from __future__ import annotations

import difflib
import re
from io import BytesIO

from openpyxl import load_workbook
from pypdf import PdfReader


# Mic-color words as they appear in 'Mic Assign' -> tag spellings Dave
# uses in channel names ("Wht HH", "Orng RF").
COLOR_SYNONYMS = {
    "red": ("red",),
    "white": ("white", "wht"),
    "blue": ("blue", "blu"),
    "black": ("black", "blk"),
    "orange": ("orange", "orng"),
    "yellow": ("yellow", "ylo"),
    "green": ("green", "grn"),
}

PAIRED_SLOTS = (1, 2, 3, 4, 5, 6)
MIC_ONLY_SLOTS = (7, 8, 9, 10)


def _s(v) -> str:
    """Cell value as a clean string. Dave's sheets have literal 0s in
    blank-ish cells; treat those as empty."""
    if v is None or v == 0:
        return ""
    return str(v).strip()


def parse_workbook(data: bytes) -> dict:
    wb = load_workbook(BytesIO(data), data_only=True)
    warnings = []

    if "Input List" not in wb.sheetnames:
        raise ValueError("Workbook has no 'Input List' sheet — is this the right file?")
    ws = wb["Input List"]

    sunday = _s(ws["A2"].value).removeprefix("Sunday:").strip() or None

    # Rows keyed by the Mic column (D): "... HH" = vocalist handheld,
    # "... RF" = wireless misc (host mics, sermon headset).
    vocals, misc = [], []
    for row in ws.iter_rows(min_row=4):
        instrument = _s(row[2].value)   # C
        mic = _s(row[3].value)          # D
        mymix = _s(row[6].value)        # G
        info = _s(row[7].value) if len(row) > 7 else ""  # H
        foh = row[1].value              # B
        entry = {
            "foh": foh if isinstance(foh, (int, float)) else None,
            "instrument": instrument,
            "mic": mic,
            "mymix": mymix,
            "info": info,
        }
        if mic.upper().endswith("HH") and instrument:
            vocals.append(entry)
        elif "RF" in mic.upper() and instrument:
            misc.append(entry)
    vocals.sort(key=lambda e: (e["foh"] is None, e["foh"]))

    # 'Mic Assign' adds stage position + IEM pack per vocalist.
    assign = {}
    if "Mic Assign" in wb.sheetnames:
        for row in wb["Mic Assign"].iter_rows(min_row=3):
            name = _s(row[0].value)
            if not name:
                continue
            assign[_person_key(name)] = {
                "position": _s(row[1].value),
                "color": _s(row[2].value),
                "pack": _s(row[3].value) if len(row) > 3 else "",
            }
    else:
        warnings.append("No 'Mic Assign' sheet — positions and IEM packs unavailable.")

    return {"sunday": sunday, "vocals": vocals, "misc": misc,
            "assign": assign, "tech": _parse_tech(wb, ws, warnings),
            "warnings": warnings}


def _parse_tech(wb, input_ws, warnings: list) -> dict:
    """The backline layer: the full patch list plus the hand-typed
    islands of the MyMix sheet (per-input monitor routing, the IEM RF
    assignments, the 16-channel MyMix legend, mixer ownership)."""
    # Per-row monitor routing from the MyMix sheet, keyed by row number
    # (its A/B/C columns are formula mirrors of Input List, same rows).
    routing = {}
    legend, iem_rf, mixers = [], [], []
    if "MyMix" in wb.sheetnames:
        mm = wb["MyMix"]
        in_iem_block = False
        for row in mm.iter_rows(min_row=4):
            r = row[0].row
            num = row[7].value if len(row) > 7 else None      # H: MyMix input #
            route = _s(row[8].value) if len(row) > 8 else ""  # I: Mtx n / Dir
            if isinstance(num, (int, float)) or route:
                routing[r] = {"num": int(num) if isinstance(num, (int, float)) else None,
                              "route": route or None}
            k = row[10].value if len(row) > 10 else None      # K column
            if isinstance(k, (int, float)) and 1 <= int(k) <= 16:
                legend.append({"ch": int(k),
                               "label": _s(row[11].value),
                               "source": _s(row[12].value)})
            elif _s(k).upper().startswith("IEM"):
                if "ASSIGNMENT" in _s(k).upper():
                    in_iem_block = True
                else:
                    iem_rf.append({"iem": _s(k), "rf": _s(row[11].value),
                                   "path": _s(row[12].value),
                                   "owner": _s(row[14].value) if len(row) > 14 else ""})
            if not in_iem_block and len(row) > 14 and _s(row[14].value):
                mixers.append({"mixer": _s(row[14].value),
                               "owner": _s(row[15].value) if len(row) > 15 else ""})
    else:
        warnings.append("No 'MyMix' sheet — routing/IEM RF/legend unavailable.")

    # Full patch list off the master sheet.
    patch = []
    for row in input_ws.iter_rows(min_row=4):
        vals = [_s(c.value) for c in row[:8]]
        if not any(vals[1:]):        # skip spacer rows
            continue
        foh = row[1].value
        r = routing.get(row[0].row, {})
        patch.append({
            "snake_ch": vals[0] or None,
            "foh_ch": int(foh) if isinstance(foh, (int, float)) else None,
            "instrument": vals[2] or None,
            "mic": vals[3] or None,
            "phantom": 1 if vals[4].upper() == "X" else 0,
            "mute_grp": vals[5] or None,
            "mymix_ch": vals[6] or None,
            "mymix_num": r.get("num"),
            "mymix_route": r.get("route"),
            "info": vals[7] or None,
        })
    return {"patch": patch, "legend": legend, "iem_rf": iem_rf, "mixers": mixers}


def _person_key(name: str) -> str:
    """'Chris Vox' -> 'chris'; 'Joe B' -> 'joe b'. Strips the role suffix
    Dave appends to vocalists."""
    name = name.strip()
    if name.lower().endswith(" vox"):
        name = name[:-4]
    return name.strip().lower()


def _looks_like_typo(a: str, b: str) -> bool:
    """Close-but-not-equal names — one or two characters apart."""
    return a.lower() != b.lower() and \
        difflib.SequenceMatcher(None, a.lower(), b.lower()).ratio() >= 0.84


def _match_person(people, raw_name: str):
    """Match a sheet name against person rows. Exact display-name match
    (ignoring a trailing initial's dot) wins; then unique first-name
    match. Returns (person, warning, create_name) — at most one set.
    A name nobody matches becomes create_name (the person is added on
    apply) UNLESS it reads like a typo of someone existing, which
    warns instead of minting a duplicate."""
    key = _person_key(raw_name)
    if not key:
        return None, None, None
    for p in people:
        if p["display_name"].rstrip(".").lower() == key:
            return p, None, None
    first_matches = [
        p for p in people
        if p["display_name"].split()[0].lower() == key.split()[0]
    ]
    if len(first_matches) == 1:
        return first_matches[0], None, None
    if len(first_matches) > 1:
        return None, f"'{raw_name}': several people share that first name — pick manually.", None
    near = [p for p in people
            if _looks_like_typo(key.split()[0], p["display_name"].split()[0])]
    if near:
        return None, (f"'{raw_name}': looks like a typo of {near[0]['display_name']}"
                      f" — fix the sheet, or add them yourself if they're real."), None
    name = raw_name.strip()
    if name.lower().endswith(" vox"):
        name = name[:-4].strip()
    return None, None, name


def _match_channel_by_tag(channels, tag: str, kinds: tuple):
    """Match 'Red HH' / 'Orng RF' / 'Sermon RF' against channel labels
    or color tags, restricted to the given kinds."""
    tag_l = tag.lower()
    pool = [c for c in channels if c["kind"] in kinds]
    # 1. Whole-tag match against label or the free-text capsule tag.
    for c in pool:
        if c["label"].lower() == tag_l or (c["capsule"] or "").lower() == tag_l:
            return c
    # 2. Color-word match ("Red HH" ~ "Red Handheld"; "Wht HH" ~ "White HH").
    words = re.split(r"[\s/]+", tag_l)
    for word in words:
        for color, spellings in COLOR_SYNONYMS.items():
            if word in spellings:
                hits = [
                    c for c in pool
                    if any(sp in c["label"].lower() for sp in spellings)
                    or any(sp in (c["capsule"] or "").lower() for sp in spellings)
                ]
                if len(hits) == 1:
                    return hits[0]
    # 3. Distinctive-word match ("Sermon RF" ~ "Sermon Headset").
    for word in words:
        if word in ("rf", "hh", "pack"):
            continue
        hits = [c for c in pool if word in c["label"].lower()]
        if len(hits) == 1:
            return hits[0]
    return None


def _match_iem(channels, pack: str):
    """'Pack 1' -> the IEM channel whose label ENDS with that number.
    The last number is the pack number — labels like 'IEM 900 Pack 1'
    carry the PSM series number first."""
    digits = re.findall(r"\d+", pack)
    if not digits:
        return None
    n = digits[-1]
    pool = [c for c in channels if c["kind"] == "iem"]
    for c in pool:
        if c["label"].lower() == pack.lower():
            return c
    hits = [c for c in pool
            if (re.findall(r"\d+", c["label"]) or [None])[-1] == n]
    if len(hits) == 1:
        return hits[0]
    return None


def build_plan(db, parsed: dict) -> dict:
    people = db.execute(
        "SELECT id, display_name FROM person WHERE archived=0"
    ).fetchall()
    channels = db.execute(
        "SELECT id, label, kind, capsule FROM channel WHERE archived=0"
    ).fetchall()
    positions = {
        r["label"].lower(): r
        for r in db.execute("SELECT id, label FROM position WHERE archived=0")
    }
    slots = {
        r["bank_order"]: r
        for r in db.execute("SELECT id, bank_order, kind FROM slot WHERE archived=0")
    }

    plan = {"sunday": parsed["sunday"], "assignments": [],
            "tech": parsed.get("tech") or {},
            "warnings": list(parsed["warnings"])}

    def resolve(entry, bank_order, mic_kinds):
        slot = slots.get(bank_order)
        if slot is None:
            return None
        a = {"slot_id": slot["id"], "bank_order": bank_order,
             "person_id": None, "person": "", "mic_channel_id": None, "mic": "",
             "iem_channel_id": None, "iem": "", "position_id": None, "position": "",
             "mymix": None, "clear": entry is None, "notes": []}
        if entry is None:
            a["notes"].append("No row this week — slot will be emptied.")
            return a

        person, warn, create = _match_person(people, entry["instrument"])
        if person is not None:
            a["person_id"], a["person"] = person["id"], person["display_name"]
        elif create:
            a["person_create"], a["person"] = create, f"{create} (new)"
            a["notes"].append(f"'{create}' will be added to People.")
        elif warn:
            a["notes"].append(warn)

        if entry["mic"]:
            ch = _match_channel_by_tag(channels, entry["mic"], mic_kinds)
            if ch is not None:
                a["mic_channel_id"], a["mic"] = ch["id"], ch["label"]
            else:
                a["notes"].append(
                    f"Mic '{entry['mic']}': no matching channel in inventory.")

        a["mymix"] = entry["mymix"] or None

        extra = parsed["assign"].get(_person_key(entry["instrument"]), {})
        if extra.get("pack") and slot["kind"] == "paired":
            iem = _match_iem(channels, extra["pack"])
            if iem is not None:
                a["iem_channel_id"], a["iem"] = iem["id"], iem["label"]
            else:
                a["notes"].append(
                    f"IEM '{extra['pack']}': no matching IEM channel in inventory.")
        if extra.get("position"):
            pos = positions.get(extra["position"].lower())
            if pos is not None:
                a["position_id"], a["position"] = pos["id"], pos["label"]
            else:
                a["position"] = extra["position"]
                a["notes"].append(f"Position '{extra['position']}' will be created.")
        return a

    vocals, misc = parsed["vocals"], parsed["misc"]
    for i, bank_order in enumerate(PAIRED_SLOTS):
        a = resolve(vocals[i] if i < len(vocals) else None, bank_order, ("handheld",))
        if a:
            plan["assignments"].append(a)
    for i, bank_order in enumerate(MIC_ONLY_SLOTS):
        a = resolve(misc[i] if i < len(misc) else None, bank_order,
                    ("beltpack", "handheld"))
        if a:
            plan["assignments"].append(a)

    if len(vocals) > len(PAIRED_SLOTS):
        plan["warnings"].append(
            f"{len(vocals) - len(PAIRED_SLOTS)} vocalist row(s) beyond slot 6 ignored.")
    if len(misc) > len(MIC_ONLY_SLOTS):
        plan["warnings"].append(
            f"{len(misc) - len(MIC_ONLY_SLOTS)} wireless row(s) beyond slot 10 ignored.")
    return plan


def _ensure_person(db, name: str, cache: dict) -> int:
    """Find-or-create a person by display name (case-insensitive).
    The cache keeps one import from creating the same name twice."""
    key = name.lower()
    if key not in cache:
        row = db.execute(
            "SELECT id FROM person WHERE archived=0 AND lower(display_name)=?",
            (key,)).fetchone()
        cache[key] = row["id"] if row else db.execute(
            "INSERT INTO person (display_name) VALUES (?)", (name,)).lastrowid
    return cache[key]


def apply_plan(db, plan: dict) -> int:
    """Write the assignments and replace the backline tech tables.
    Creates any positions and people the plan flagged. Returns slots
    updated."""
    assignments = plan.get("assignments", [])
    created: dict = {}
    for a in assignments:
        if a.get("person_id") is None and a.get("person_create"):
            a["person_id"] = _ensure_person(db, a["person_create"], created)
        if a["position_id"] is None and a["position"]:
            row = db.execute(
                "SELECT id FROM position WHERE archived=0 AND lower(label)=lower(?)",
                (a["position"],),
            ).fetchone()
            if row is None:
                cur = db.execute("INSERT INTO position (label) VALUES (?)",
                                 (a["position"],))
                a["position_id"] = cur.lastrowid
            else:
                a["position_id"] = row["id"]
        # pc_role is cleared here and re-stamped by the Tech Report
        # import (step 2) — stale roles must not survive a people swap.
        db.execute(
            """UPDATE slot SET
                 person_id=?, mic_channel_id=?, iem_channel_id=?,
                 position_id=?, mymix_channel=?, pc_role=NULL,
                 updated_at=datetime('now')
               WHERE id=?""",
            (a["person_id"], a["mic_channel_id"], a["iem_channel_id"],
             a["position_id"], a["mymix"], a["slot_id"]),
        )

    # Backline tech tables are weekly truth: full replace on each apply.
    tech = plan.get("tech") or {}
    if tech:
        db.execute("DELETE FROM patch_row")
        for i, p in enumerate(tech.get("patch", [])):
            db.execute(
                """INSERT INTO patch_row (sort_order, snake_ch, foh_ch, instrument,
                     mic, phantom, mute_grp, mymix_ch, mymix_num, mymix_route, info)
                   VALUES (?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?)""",
                (i, p.get("snake_ch"), p.get("foh_ch"), p.get("instrument"),
                 p.get("mic"), p.get("phantom", 0), p.get("mute_grp"),
                 p.get("mymix_ch"), p.get("mymix_num"), p.get("mymix_route"),
                 p.get("info")))
        db.execute("DELETE FROM iem_rf")
        for i, r in enumerate(tech.get("iem_rf", [])):
            db.execute("INSERT INTO iem_rf (sort_order, iem, rf, path, owner) "
                       "VALUES (?, ?, ?, ?, ?)",
                       (i, r.get("iem"), r.get("rf") or None,
                        r.get("path") or None, r.get("owner") or None))
        db.execute("DELETE FROM mymix_channel")
        for l in tech.get("legend", []):
            db.execute("INSERT INTO mymix_channel (ch, label, source) VALUES (?, ?, ?)",
                       (l["ch"], l.get("label") or None, l.get("source") or None))
        db.execute("DELETE FROM mymix_mixer")
        for i, m in enumerate(tech.get("mixers", [])):
            db.execute("INSERT INTO mymix_mixer (sort_order, mixer, owner) VALUES (?, ?, ?)",
                       (i, m["mixer"], m.get("owner") or None))
        for key, value in (("import_source", plan.get("sunday")),
                           ("import_at", None)):
            db.execute(
                """INSERT INTO app_setting (key, value)
                   VALUES (?, COALESCE(?, datetime('now')))
                   ON CONFLICT(key) DO UPDATE SET value=excluded.value""",
                (key, value))
    db.commit()
    return len(assignments)


# =================================================================
# Tech Report (Planning Center service-plan PDF) importer
# =================================================================
# Dave prints this report weekly. Page 2 carries the plan positions:
# tech crew, band, music leader, host, speaker, vocals, stream crew.
# Same two-step flow as the workbook: parse -> plan -> apply.

TECH_POSITIONS = {
    "backstage manager", "graphics", "lights", "service director", "sound",
    "acoustic & vocals", "bass guitar", "drums", "electric guitar",
    "keys", "keys & vocals", "synth", "vocals", "speaker", "host",
    "sign language interpreter", "director/switcher",
}
TECH_CATEGORIES = {
    "audio/visual", "band", "interpreters", "live stream",
    "music leader", "preacher/speaker", "service host", "vocals",
}
# Position -> tech seat (bank_order), matched loosely.
TECH_SEAT_MAP = [
    (r"^sound$", 17), (r"backstage", 18), (r"service director", 19),
    (r"graphic", 20), (r"^lights?$", 21), (r"switcher|stream", 22),
]
# Positions that imply a band seat, matched against seat labels.
BAND_POS_MAP = [
    (r"drum", r"drum"), (r"bass", r"bass"), (r"electric", r"elec"),
    (r"acoustic", r"acou"), (r"keys", r"key"), (r"synth", r"synth"),
]
_NAME_RE = re.compile(r"^[A-Z][\w.'-]*(?: [A-Z][\w.'-]*)+(?: \?)?$")


def parse_tech_report(data: bytes) -> dict:
    reader = PdfReader(BytesIO(data))
    lines = []
    for page in reader.pages:
        lines.extend((page.extract_text() or "").splitlines())

    date = None
    for ln in lines:
        m = re.search(r"([A-Z][a-z]+ \d{1,2}, \d{4})", ln)
        if m:
            date = m.group(1)
            break

    # The service rundown (page 1) reuses words like 'Host' as section
    # headers, so only harvest from the assignments region — it always
    # opens with the 'Audio/Visual' category. Fall back to the whole
    # document if the marker is ever missing.
    for i, ln in enumerate(lines):
        if ln.strip().lower() == "audio/visual":
            lines = lines[i:]
            break

    # Walk lines: known position names open a position, category headers
    # set context, capitalized name-shaped lines are the open position's
    # people. 'Vocals' is both category and position — checking position
    # first makes the doubled line self-resolving (the first opens an
    # empty position, the second reopens it and collects the names).
    positions, current, category = [], None, None
    leader_pending, leader_name = False, None
    for raw in lines:
        ln = raw.strip()
        low = ln.lower()
        if not ln:
            continue
        if low in TECH_POSITIONS or re.match(r"^camera \d+$", low):
            current = {"category": category, "position": ln, "people": []}
            positions.append(current)
            continue
        if low in TECH_CATEGORIES:
            category = ln
            leader_pending = low == "music leader"
            current = None
            continue
        if current is not None and _NAME_RE.match(ln) and not any(c.isdigit() for c in ln):
            # Some report layouts interleave the call-time column with
            # the assignments — schedule vocabulary is never a name.
            SCHED_WORDS = {"call", "reh", "rehearsal", "through", "doors",
                           "open", "singing", "worship", "stream", "run",
                           "soundcheck", "time", "band", "tech"}
            if any(w.lower() in SCHED_WORDS for w in ln.split()):
                continue
            name = ln.rstrip(" ?").strip()
            current["people"].append({"name": name, "tentative": ln.endswith("?")})
            if leader_pending:
                leader_name = name
                leader_pending = False
    return {"date": date, "positions": positions, "leader_name": leader_name}


def _pc_person_match(people, full_name: str):
    """Match a PC full name against our display names: first name plus
    last initial, EXACT only. PC always prints full names, so a loose
    first-name fallback is how 'Dave Geer' once became Dave H.

    Returns (person, warning, create_name) — at most one set. An
    unknown name becomes create_name ('Dave Geer' -> 'Dave G.', added
    on apply) UNLESS it reads like a typo of an existing person with
    the same last initial, which warns instead: a DIFFERENT last
    initial is a different human, a garbled first name over the same
    initial is probably a misspelling."""
    parts = full_name.split()
    display = f"{parts[0]} {parts[-1][0]}."
    for p in people:
        if p["display_name"].lower() == display.lower():
            return p, None, None
    for p in people:
        pp = p["display_name"].split()
        if (len(pp) > 1 and pp[-1].rstrip(".").lower() == parts[-1][0].lower()
                and _looks_like_typo(parts[0], pp[0])):
            return None, (f"'{full_name}': looks like a typo of {p['display_name']}"
                          f" — fix it in Planning Center, or add them yourself "
                          f"if they're real."), None
    return None, None, display


def build_tech_plan(db, parsed: dict) -> dict:
    people = db.execute("SELECT id, display_name FROM person WHERE archived=0").fetchall()
    band_slots = db.execute(
        "SELECT id, bank_order, label FROM slot WHERE kind='band' AND archived=0").fetchall()

    plan = {"date": parsed.get("date"), "leader": None, "tech_seats": [],
            "band": [], "roles": [], "info": [], "warnings": []}

    for entry in parsed["positions"]:
        pos = entry["position"]
        low = pos.lower()
        for who in entry["people"]:
            person, warn, create = _pc_person_match(people, who["name"])
            if warn:
                plan["warnings"].append(f"{pos}: {warn}")
                continue
            # Either a matched row or a to-be-created name; downstream
            # entries carry person_create so apply can mint the person.
            pid = person["id"] if person else None
            pname = person["display_name"] if person else f"{create} (new)"
            ref = {"person_id": pid, "person": pname}
            if create:
                ref["person_create"] = create

            seat = next((s for rx, s in TECH_SEAT_MAP if re.search(rx, low)), None)
            if seat is not None and "vocal" not in low:
                plan["tech_seats"].append({"seat": seat, "role": pos, **ref})
                continue

            if re.match(r"^camera \d+$", low) or "interpreter" in low:
                plan["info"].append(f"{pos}: {pname}")
                if create:
                    plan["extra_people"] = plan.get("extra_people", []) + [create]
                continue

            band_rx = next((brx for prx, brx in BAND_POS_MAP if re.search(prx, low)), None)
            if band_rx:
                slot = next((s for s in band_slots
                             if re.search(band_rx, (s["label"] or ""), re.I)), None)
                if slot:
                    plan["band"].append({"bank_order": slot["bank_order"],
                                         "seat_label": slot["label"],
                                         "position": pos, **ref})
                else:
                    plan["warnings"].append(f"{pos}: no band seat matches.")
            plan["roles"].append({"role": pos, **ref})

    if parsed.get("leader_name"):
        person, warn, create = _pc_person_match(people, parsed["leader_name"])
        if person:
            plan["leader"] = {"person_id": person["id"], "person": person["display_name"]}
        elif create:
            plan["leader"] = {"person_id": None, "person": f"{create} (new)",
                              "person_create": create}
        elif warn:
            plan["warnings"].append(f"Music leader: {warn}")
    return plan


def apply_tech_plan(db, plan: dict) -> int:
    """Write tech seats, band assignments, weekly PC roles, the leader.

    Starts by clearing everything this import owns — tech seats and
    band seats — so last week's people can never linger on a seat the
    new plan doesn't mention (the Dave-on-two-seats bug)."""
    db.execute("UPDATE slot SET person_id=NULL, pc_role=NULL, "
               "updated_at=datetime('now') WHERE kind='tech' AND archived=0")
    db.execute("UPDATE slot SET person_id=NULL, pc_role=NULL, iem_channel_id=NULL, "
               "mymix_channel=NULL, updated_at=datetime('now') "
               "WHERE kind='band' AND archived=0")
    # Mint any new people the plan carries (find-or-create; one row per
    # name even if they hold several positions).
    created: dict = {}
    for entry in (plan.get("tech_seats", []) + plan.get("band", [])
                  + plan.get("roles", [])
                  + ([plan["leader"]] if plan.get("leader") else [])):
        if entry.get("person_id") is None and entry.get("person_create"):
            entry["person_id"] = _ensure_person(db, entry["person_create"], created)
    for name in plan.get("extra_people", []):
        _ensure_person(db, name, created)
    changed = 0
    for t in plan.get("tech_seats", []):
        db.execute("UPDATE slot SET person_id=?, pc_role=?, updated_at=datetime('now') "
                   "WHERE bank_order=? AND kind='tech'",
                   (t["person_id"], t["role"], t["seat"]))
        changed += 1
    for b in plan.get("band", []):
        db.execute("UPDATE slot SET person_id=?, pc_role=?, updated_at=datetime('now') "
                   "WHERE bank_order=?", (b["person_id"], b["position"], b["bank_order"]))
        changed += 1
    for r in plan.get("roles", []):
        db.execute("UPDATE slot SET pc_role=?, updated_at=datetime('now') "
                   "WHERE person_id=? AND kind != 'tech'", (r["role"], r["person_id"]))
    if plan.get("leader"):
        db.execute("INSERT INTO app_setting (key, value) VALUES ('leader_person_id', ?) "
                   "ON CONFLICT(key) DO UPDATE SET value=excluded.value",
                   (str(plan["leader"]["person_id"]),))
    for key, value in (("tech_report_source", plan.get("date")), ("tech_report_at", None)):
        db.execute("INSERT INTO app_setting (key, value) "
                   "VALUES (?, COALESCE(?, datetime('now'))) "
                   "ON CONFLICT(key) DO UPDATE SET value=excluded.value", (key, value))
    db.commit()
    return changed
